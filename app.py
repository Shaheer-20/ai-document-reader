import os
import time
import fitz  # PyMuPDF
import requests
import csv
import io
import openpyxl
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, jsonify, session, send_file
from dotenv import load_dotenv
import uuid
import base64
import json
from datetime import datetime
import sqlite3
import hashlib

# Load environment variables from a .env file
load_dotenv()

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'

# --- Database Configuration ---
DATABASE = 'documents.db'

def init_db():
    """Initialize the database with required tables."""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # Create documents table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            file_hash TEXT UNIQUE NOT NULL,
            file_size INTEGER NOT NULL,
            file_type TEXT NOT NULL,
            upload_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            text_content TEXT,
            summary TEXT,
            preview_data TEXT,
            session_id TEXT
        )
    ''')
    
    # Create chat_history table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS chat_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER,
            question TEXT NOT NULL,
            answer TEXT NOT NULL,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (document_id) REFERENCES documents (id)
        )
    ''')
    
    conn.commit()
    conn.close()

def get_db_connection():
    """Get database connection."""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def store_document(file_data, filename, text_content, summary=None, preview_data=None, session_id=None):
    """Store document in database."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Generate file hash for uniqueness
    file_hash = hashlib.md5(file_data).hexdigest()
    file_size = len(file_data)
    file_type = os.path.splitext(filename)[1].lower()
    
    try:
        cursor.execute('''
            INSERT INTO documents (filename, file_hash, file_size, file_type, text_content, summary, preview_data, session_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (filename, file_hash, file_size, file_type, text_content, summary, json.dumps(preview_data) if preview_data else None, session_id))
        
        document_id = cursor.lastrowid
        conn.commit()
        print(f"Document stored successfully: ID={document_id}, Filename={filename}")  # Debug log
        return document_id
    except sqlite3.IntegrityError:
        # Document already exists, get existing ID
        cursor.execute('SELECT id FROM documents WHERE file_hash = ?', (file_hash,))
        result = cursor.fetchone()
        document_id = result['id'] if result else None
        conn.commit()
        print(f"Document already exists: ID={document_id}, Filename={filename}")  # Debug log
        return document_id
    except Exception as e:
        print(f"Error storing document: {e}")  # Debug log
        conn.rollback()
        return None
    finally:
        conn.close()

def get_document_by_id(document_id):
    """Get document by ID."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM documents WHERE id = ?', (document_id,))
    document = cursor.fetchone()
    conn.close()
    return document

def get_documents_by_session(session_id):
    """Get all documents for a session."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM documents WHERE session_id = ? ORDER BY upload_time DESC', (session_id,))
    documents = cursor.fetchall()
    conn.close()
    print(f"Retrieved {len(documents)} documents for session {session_id}")  # Debug log
    return documents

def store_chat_message(document_id, question, answer):
    """Store chat message in database."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO chat_history (document_id, question, answer)
        VALUES (?, ?, ?)
    ''', (document_id, question, answer))
    conn.commit()
    conn.close()

def get_chat_history(document_id):
    """Get chat history for a document."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM chat_history WHERE document_id = ? ORDER BY timestamp ASC', (document_id,))
    history = cursor.fetchall()
    conn.close()
    return history

# Initialize database on startup
init_db()

# --- Gemini API Configuration ---
API_KEY = os.getenv("GEMINI_API_KEY")
GEMINI_API_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-preview-09-2025:generateContent?key={API_KEY}"
ALLOWED_EXTENSIONS = {'.pdf', '.csv', '.xlsx', '.html'}

# --- Text Extraction Functions ---

def extract_text_from_pdf(stream):
    """Extracts text from a PDF file stream."""
    doc = fitz.open(stream=stream, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text

def extract_text_from_excel(stream):
    """Extracts text from an XLSX (Excel) file stream."""
    text = ""
    workbook = openpyxl.load_workbook(stream)
    for sheet in workbook.worksheets:
        text += f"Sheet: {sheet.title}\n"
        for row in sheet.iter_rows():
            row_text = [str(cell.value) if cell.value is not None else "" for cell in row]
            text += ", ".join(row_text) + "\n"
    return text

def extract_text_from_csv(stream):
    """Extracts text from a CSV file stream."""
    text = ""
    # We need to decode the byte stream into text
    try:
        data = stream.read().decode('utf-8')
    except UnicodeDecodeError:
        data = stream.read().decode('latin-1') # Fallback encoding
        
    reader = csv.reader(io.StringIO(data))
    for row in reader:
        text += ", ".join(row) + "\n"
    return text

def extract_text_from_html(stream):
    """Extracts visible text from an HTML file stream."""
    soup = BeautifulSoup(stream, "html.parser")
    # Get all visible text
    text = soup.get_text(separator="\n", strip=True)
    return text

def extract_text_from_file(file_stream, filename):
    """Main function to route to the correct text extractor."""
    try:
        file_ext = os.path.splitext(filename)[1].lower()

        if file_ext == '.pdf':
            return extract_text_from_pdf(file_stream)
        elif file_ext == '.xlsx':
            return extract_text_from_excel(file_stream)
        elif file_ext == '.csv':
            return extract_text_from_csv(file_stream)
        elif file_ext == '.html':
            return extract_text_from_html(file_stream)
        else:
            return None
    except Exception as e:
        print(f"Error extracting text from {filename}: {e}")
        return None

def generate_document_preview(file_stream, filename):
    """Generate preview for different document types."""
    try:
        file_ext = os.path.splitext(filename)[1].lower()
        
        if file_ext == '.pdf':
            return generate_pdf_preview(file_stream)
        elif file_ext == '.html':
            return generate_html_preview(file_stream)
        elif file_ext in ['.csv', '.xlsx']:
            return generate_table_preview(file_stream, file_ext)
        else:
            return None
    except Exception as e:
        print(f"Error generating preview for {filename}: {e}")
        return None

def generate_pdf_preview(file_stream):
    """Generate PDF preview as base64 images."""
    try:
        doc = fitz.open(stream=file_stream, filetype="pdf")
        preview_pages = []
        total_pages = len(doc)
        
        # Get first 3 pages for preview with optimized settings
        for page_num in range(min(3, total_pages)):
            page = doc[page_num]
            # Use smaller matrix for faster processing
            pix = page.get_pixmap(matrix=fitz.Matrix(0.4, 0.4), alpha=False)
            img_data = pix.tobytes("png")
            img_base64 = base64.b64encode(img_data).decode()
            preview_pages.append({
                'page': page_num + 1,
                'image': img_base64,
                'width': pix.width,
                'height': pix.height
            })
            # Clean up pixmap to free memory
            pix = None
        
        doc.close()
        return {
            'type': 'pdf',
            'pages': preview_pages,
            'total_pages': total_pages
        }
    except Exception as e:
        print(f"Error generating PDF preview: {e}")
        return None

def generate_html_preview(file_stream):
    """Generate HTML preview."""
    try:
        soup = BeautifulSoup(file_stream, "html.parser")
        # Get first 500 characters of visible text
        text = soup.get_text(separator=" ", strip=True)[:500]
        return {
            'type': 'html',
            'preview_text': text,
            'title': soup.title.string if soup.title else 'HTML Document'
        }
    except Exception as e:
        print(f"Error generating HTML preview: {e}")
        return None

def generate_table_preview(file_stream, file_ext):
    """Generate table preview for CSV/Excel files."""
    try:
        if file_ext == '.csv':
            data = file_stream.read().decode('utf-8')
            reader = csv.reader(io.StringIO(data))
            rows = list(reader)[:10]  # First 10 rows
        else:  # .xlsx
            workbook = openpyxl.load_workbook(file_stream)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(max_row=10, values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
        
        return {
            'type': 'table',
            'data': rows,
            'file_type': file_ext
        }
    except Exception as e:
        print(f"Error generating table preview: {e}")
        return None

# --- Gemini API Function ---

def get_summary_from_gemini(text_content):
    """
    Calls the Gemini API to summarize the given text.
    Implements exponential backoff for retries.
    """
    if not API_KEY:
        return None, "Error: GEMINI_API_KEY is not set in the environment."

    # Truncate text if it's too long for a single API request (adjust as needed)
    # A simple way to avoid hitting request size limits.
    max_chars = 100000 
    if len(text_content) > max_chars:
        text_content = text_content[:max_chars] + "\n... [Content Truncated]"

    payload = {
        "contents": [
            {
                "parts": [
                    {"text": f"Please provide a concise, multi-paragraph summary of the following document content:\n\n---\n\n{text_content}"}
                ]
            }
        ],
        "generationConfig": {
            "temperature": 0.5,
            "topK": 1,
            "topP": 1,
            "maxOutputTokens": 2048,
        }
    }

    headers = {
        'Content-Type': 'application/json'
    }

    max_retries = 3
    delay = 1  # Initial delay in seconds

    for attempt in range(max_retries):
        try:
            response = requests.post(GEMINI_API_URL, headers=headers, json=payload, timeout=60)

            # Successful request
            if response.status_code == 200:
                result = response.json()
                if 'candidates' in result:
                    summary = result['candidates'][0]['content']['parts'][0]['text']
                    return summary, None
                else:
                    return None, "Error: Invalid response from API."

            # Rate limiting or server overload
            elif response.status_code in [429, 503]:
                print(f"API rate limit or server error. Retrying in {delay}s...")
                time.sleep(delay)
                delay *= 2  # Exponential backoff
            
            # Other client/server errors
            else:
                error_details = response.text
                print(f"API Error: {response.status_code}\n{error_details}")
                return None, f"API Error: {response.status_code}. Check server logs."

        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}. Retrying in {delay}s...")
            time.sleep(delay)
            delay *= 2

    return None, "Error: API request failed after all retries."

def get_chat_response_from_gemini(question, document_text):
    """
    Calls the Gemini API to answer questions about the document.
    """
    if not API_KEY:
        return None, "Error: GEMINI_API_KEY is not set in the environment."

    # Truncate text if it's too long for a single API request
    max_chars = 100000 
    if len(document_text) > max_chars:
        document_text = document_text[:max_chars] + "\n... [Content Truncated]"

    payload = {
        "contents": [
            {
                "parts": [
                    {"text": f"Based on the following document content, please answer this question: {question}\n\nDocument content:\n---\n{document_text}\n\nPlease provide a detailed and accurate answer based only on the information in the document. If the answer cannot be found in the document, please say so."}
                ]
            }
        ],
        "generationConfig": {
            "temperature": 0.3,
            "topK": 1,
            "topP": 1,
            "maxOutputTokens": 1024,
        }
    }

    headers = {
        'Content-Type': 'application/json'
    }

    max_retries = 3
    delay = 1

    for attempt in range(max_retries):
        try:
            response = requests.post(GEMINI_API_URL, headers=headers, json=payload, timeout=60)

            if response.status_code == 200:
                result = response.json()
                if 'candidates' in result:
                    answer = result['candidates'][0]['content']['parts'][0]['text']
                    return answer, None
                else:
                    return None, "Error: Invalid response from API."

            elif response.status_code in [429, 503]:
                print(f"API rate limit or server error. Retrying in {delay}s...")
                time.sleep(delay)
                delay *= 2
            
            else:
                error_details = response.text
                print(f"API Error: {response.status_code}\n{error_details}")
                return None, f"API Error: {response.status_code}. Check server logs."

        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}. Retrying in {delay}s...")
            time.sleep(delay)
            delay *= 2

    return None, "Error: API request failed after all retries."

# --- Flask Routes ---

@app.route('/', methods=['GET', 'POST'])
def index():
    summary = None
    error = None
    document_uploaded = False

    if request.method == 'POST':
        # 1. Check if the file part is in the request
        if 'pdf_file' not in request.files:
            error = "No file part in the request. Please select a file."
            return render_template('index.html', summary=summary, error=error, document_uploaded=document_uploaded)

        file = request.files['pdf_file']

        # 2. Check if a file was actually selected
        if file.filename == '':
            error = "No file selected. Please choose a document to upload."
            return render_template('index.html', summary=summary, error=error, document_uploaded=document_uploaded)

        # 3. Check if the file type is allowed
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file and file_ext in ALLOWED_EXTENSIONS:
            try:
                # Read the file into memory
                file_stream = file.read()

                # 4. Extract text
                text = extract_text_from_file(io.BytesIO(file_stream), file.filename)

                if text:
                    print(f"Text extracted successfully, length: {len(text)}")  # Debug log
                    
                    # Get summary from Gemini instantly
                    summary, error = get_summary_from_gemini(text)
                    
                    if error:
                        print(f"Gemini API error: {error}")  # Debug log
                        error = f"Failed to generate summary: {error}"
                    elif summary:
                        print(f"Summary generated successfully, length: {len(summary)}")  # Debug log
                        
                        # Store document in database
                        session_id = session.get('session_id', str(uuid.uuid4()))
                        if 'session_id' not in session:
                            session['session_id'] = session_id
                        
                        # Store document without preview data for now
                        document_id = store_document(
                            file_stream, 
                            file.filename, 
                            text, 
                            summary, 
                            None,  # No preview data
                            session_id
                        )
                        
                        # Store document info in session for chat functionality
                        session['document_id'] = document_id
                        session['document_text'] = text
                        session['document_filename'] = file.filename
                        session['document_size'] = len(file_stream)
                        session['upload_time'] = datetime.now().isoformat()
                        document_uploaded = True
                        
                        print(f"Document uploaded and stored successfully: {file.filename}")  # Debug log
                    else:
                        print("No summary generated")  # Debug log
                        error = "Failed to generate summary from the document."
                else:
                    error = f"Could not extract any text from the file. The file might be empty, corrupted, or an unsupported format."
            
            except Exception as e:
                print(f"An error occurred: {e}")
                error = "An unexpected error occurred during processing."
        
        else:
            error = f"Invalid file type. Allowed types are: {', '.join(ALLOWED_EXTENSIONS)}"

    return render_template('index.html', summary=summary, error=error, document_uploaded=document_uploaded)

@app.route('/chat', methods=['POST'])
def chat():
    """Handle chat questions about the uploaded document."""
    try:
        if 'document_text' not in session:
            return jsonify({'error': 'No document uploaded. Please upload a document first.'}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid request data.'}), 400
            
        question = data.get('question', '').strip()
        
        if not question:
            return jsonify({'error': 'Please provide a question.'}), 400
        
        print(f"Chat request: {question}")  # Debug log
        
        # Get the answer from Gemini
        answer, error = get_chat_response_from_gemini(question, session['document_text'])
        
        if error:
            print(f"Gemini API error: {error}")  # Debug log
            return jsonify({'error': error}), 500
        
        # Store chat message in database
        if 'document_id' in session:
            try:
                store_chat_message(session['document_id'], question, answer)
            except Exception as db_error:
                print(f"Database error: {db_error}")  # Debug log
                # Don't fail the request if database storage fails
        
        return jsonify({'answer': answer})
        
    except Exception as e:
        print(f"Chat endpoint error: {e}")  # Debug log
        return jsonify({'error': 'An unexpected error occurred.'}), 500

# Preview functionality removed - focusing on core functionality

@app.route('/document-info')
def document_info():
    """Get document information."""
    if 'document_filename' not in session:
        return jsonify({'error': 'No document uploaded'}), 400
    
    return jsonify({
        'filename': session.get('document_filename'),
        'size': session.get('document_size'),
        'upload_time': session.get('upload_time'),
        'preview_type': session.get('document_preview', {}).get('type')
    })

@app.route('/export-summary')
def export_summary():
    """Export summary as text file."""
    if 'document_text' not in session:
        return jsonify({'error': 'No document uploaded'}), 400
    
    summary = request.args.get('summary', '')
    filename = session.get('document_filename', 'document')
    name_without_ext = os.path.splitext(filename)[0]
    
    # Create a simple text file
    output = io.StringIO()
    output.write(f"Document Summary for: {filename}\n")
    output.write("=" * 50 + "\n\n")
    output.write(summary)
    output.write("\n\n" + "=" * 50 + "\n")
    output.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        as_attachment=True,
        download_name=f"{name_without_ext}_summary.txt",
        mimetype='text/plain'
    )

@app.route('/clear-session', methods=['POST'])
def clear_session():
    """Clear the current session."""
    session.clear()
    return jsonify({'success': True})

@app.route('/test-api', methods=['GET'])
def test_api():
    """Test if the Gemini API is working."""
    if not API_KEY:
        return jsonify({'error': 'GEMINI_API_KEY is not set'}), 400
    
    try:
        # Test with a simple request
        test_payload = {
            "contents": [
                {
                    "parts": [
                        {"text": "Hello, this is a test message. Please respond with 'API is working'."}
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.1,
                "topK": 1,
                "topP": 1,
                "maxOutputTokens": 50,
            }
        }
        
        headers = {'Content-Type': 'application/json'}
        response = requests.post(GEMINI_API_URL, headers=headers, json=test_payload, timeout=10)
        
        if response.status_code == 200:
            result = response.json()
            if 'candidates' in result:
                return jsonify({'success': True, 'message': 'API is working correctly'})
            else:
                return jsonify({'error': 'Invalid API response format'}), 500
        else:
            return jsonify({'error': f'API returned status {response.status_code}: {response.text}'}), 500
            
    except Exception as e:
        return jsonify({'error': f'API test failed: {str(e)}'}), 500

@app.route('/test-db', methods=['GET'])
def test_database():
    """Test if the database is working."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Test database connection
        cursor.execute('SELECT COUNT(*) as count FROM documents')
        result = cursor.fetchone()
        total_documents = result['count']
        
        # Get session documents if session exists
        session_documents = 0
        if 'session_id' in session:
            cursor.execute('SELECT COUNT(*) as count FROM documents WHERE session_id = ?', (session['session_id'],))
            result = cursor.fetchone()
            session_documents = result['count']
        
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Database is working correctly',
            'total_documents': total_documents,
            'session_documents': session_documents,
            'session_id': session.get('session_id', 'No session')
        })
        
    except Exception as e:
        return jsonify({'error': f'Database test failed: {str(e)}'}), 500

@app.route('/documents')
def get_documents():
    """Get all documents for the current session."""
    if 'session_id' not in session:
        return jsonify({'documents': []})
    
    documents = get_documents_by_session(session['session_id'])
    document_list = []
    
    for doc in documents:
        document_list.append({
            'id': doc['id'],
            'filename': doc['filename'],
            'file_size': doc['file_size'],
            'file_type': doc['file_type'],
            'upload_time': doc['upload_time'],
            'has_summary': bool(doc['summary'])
        })
    
    return jsonify({'documents': document_list})

@app.route('/document/<int:document_id>')
def get_document(document_id):
    """Get specific document details."""
    document = get_document_by_id(document_id)
    if not document:
        return jsonify({'error': 'Document not found'}), 404
    
    return jsonify({
        'id': document['id'],
        'filename': document['filename'],
        'file_size': document['file_size'],
        'file_type': document['file_type'],
        'upload_time': document['upload_time'],
        'summary': document['summary'],
        'preview_data': json.loads(document['preview_data']) if document['preview_data'] else None
    })

@app.route('/document/<int:document_id>/chat-history')
def get_document_chat_history(document_id):
    """Get chat history for a specific document."""
    history = get_chat_history(document_id)
    chat_list = []
    
    for msg in history:
        chat_list.append({
            'question': msg['question'],
            'answer': msg['answer'],
            'timestamp': msg['timestamp']
        })
    
    return jsonify({'chat_history': chat_list})

@app.route('/document/<int:document_id>/load', methods=['POST'])
def load_document(document_id):
    """Load a specific document into the current session."""
    document = get_document_by_id(document_id)
    if not document:
        return jsonify({'error': 'Document not found'}), 404
    
    # Load document into session
    session['document_id'] = document['id']
    session['document_text'] = document['text_content']
    session['document_filename'] = document['filename']
    session['document_size'] = document['file_size']
    session['upload_time'] = document['upload_time']
    session['document_preview'] = json.loads(document['preview_data']) if document['preview_data'] else None
    
    return jsonify({'success': True, 'message': 'Document loaded successfully'})

if __name__ == '__main__':
    # Use 0.0.0.0 to make it accessible on your network, or keep 127.0.0.1
    app.run(host='127.0.0.1', port=5000, debug=True)


